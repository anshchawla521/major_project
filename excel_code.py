import openpyxl
path_to_database = "dataset.xlsx"   # enter path in raw form

dataset = openpyxl.load_workbook(path_to_database)

sheet = dataset.active

def check(uniqueid):
    i = 2
    flag = True
    while flag:
        #print(i)
        x = sheet.cell(row = i, column= 1)
        # print(x, x.value, type(x.value))
        if x.value == None:
            break
        elif int(x.value) == uniqueid:
            print("found")
            return i
        i += 1
        
    return i

def read(uniqueid):
    present = check(uniqueid)
    if sheet.cell(row = present, column = 1).value == None:
        print("no")
        return None
    else:
        send = dict()
        col = 1
        while col <= sheet.max_column:
            cell = sheet.cell(row = 1, column = col)
            temp = sheet.cell(row = present, column= col)
            if cell.value == "Student Name":
                send.update({"name":temp.value})
            elif str(cell.value) == "Student ID":
                send.update({"sid":temp.value})
            elif str(cell.value) == "Unique ID":
                send.update({"uid":int(temp.value)})
            elif cell.value == "Student Phone Number":
                send.update({"phone":temp.value})
            elif cell.value == "Location":
                send.update({"location":temp.value})
            #print(cell.value)
            col += 1
            if len(send) == 5:
                break
        return send

def write(data):
    present = check(int(data["uid"]))
    print(present)
    # if sheet.cell(row = present, column = 1).value == data["uid"]:
    #     print("present")
    #     return False
    # else:
    col = 1
    while col <= sheet.max_column:
        #print(present)
        cell = sheet.cell(row = 1, column = col)
        temp = sheet.cell(row = present, column= col)
        if cell.value == "Student Name":
            temp.value = data["name"]
        elif str(cell.value) == "Student ID":
            temp.value = data["sid"]
        elif str(cell.value) == "Unique ID":
            temp.value = int(data["uid"])
        elif cell.value == "Student Phone Number":
            temp.value = data["phone"]
        elif cell.value == "Location":
            temp.value = data["location"]
        #print(cell.value)
        col += 1
    return True



write({"name" : "Ashok", "sid" : "19105092", "uid" : 2279590476710 ,"phone":"8968963929", "location" : "in"})

#print(check(902634020041))

dataset.save("dataset.xlsx")