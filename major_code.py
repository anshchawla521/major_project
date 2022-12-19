import RPi.GPIO as GPIO
from mfrc522 import SimpleMFRC522
import datetime

from openpyxl import load_workbook
import time 
import os
import numpy as np



path_to_dataset = "./images"
path_to_finger_image = "image.jpg"


mapping = {"time":"A",
            "uid":"B",
            "name":"C",
            "sid":"D",
            "phone":"E",
            "direction":"F",
            "remarks":"G"}


import cv2, picamera

reader = SimpleMFRC522()

GOINGOUT = True
GOINGIN = False

path_to_database = "dataset.xlsx"   # enter path in raw form

dataset = load_workbook(path_to_database)

sheet = dataset.active


workbook = load_workbook(filename="attendance.xlsx")
sheet = workbook["Sheet1"]


direction = GOINGIN

led = 7
touch = 11
GPIO.setmode(GPIO.BOARD)
GPIO.setup(led, GPIO.OUT)
GPIO.output(led,GPIO.LOW)

GPIO.setup(touch, GPIO.IN)

camera = picamera.PiCamera()

def check(uniqueid):
    i = 2
    flag = True
    while flag:
        print(i)
        x = sheet.cell(row = i, column= 1)
        # print(x, x.value, type(x.value))
        if x.value == None:
            break
        elif int(x.value) == uniqueid:
            return i
        i += 1
        
    return i

def read(uniqueid):
    present = check(uniqueid)
    if sheet.cell(row = present, column = 1).value == None:
        print("no")
        return None
    else:
        send = dict()
        col = 1
        while col <= 11:
            cell = sheet.cell(row = 1, column = col)
            temp = sheet.cell(row = present, column= col)
            if cell.value == "Student Name":
                send.update({"name":temp.value})
            elif str(cell.value) == "Student ID":
                send.update({"sid":temp.value})
            elif str(cell.value) == "Unique ID":
                send.update({"uid":temp.value})
            elif cell.value == "Student Phone Number":
                send.update({"phone":temp.value})
            col += 1
            if len(send) == 4:
                break
        return send

def write(data):
    present = check(data["uid"])
    # if sheet.cell(row = present, column = 1).value == data["uid"]:
    #     print("present")
    #     return False
    # else:
    col = 1
    while col <= 11:
        print(present)
        cell = sheet.cell(row = 1, column = col)
        temp = sheet.cell(row = present, column= col)
        if cell.value == "Student Name":
            temp.value = data["name"]
        elif str(cell.value) == "Student ID":
            temp.value = data["sid"]
        elif str(cell.value) == "Unique ID":
            temp.value = data["uid"]
        elif cell.value == "Student Phone Number":
            temp.value = data["phone"]
        elif cell.value == "Location":
            temp.value = data["location"]
        col += 1
    return True



def match_finger(image)->bool:
    sample = cv2.imread(path_to_finger_image)

    # cv2.imshow("Wihout filter",sample)
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()
    sample = cv2.cvtColor(sample, cv2.COLOR_BGR2GRAY)
    sample = abs(140-sample)
    sample = sample[90:310, 170:480]
    kernel = np.array([[0, -1, 0],
                    [-1, 6, -1],
                    [0, -1, 0]])
    # kernel = np.array([[-1, -1, -1],
    #                [-1, 9,-1],
    #                [-1, -1, -1]])

    sample = cv2.filter2D(src=sample, ddepth=-1, kernel=kernel)


    #cv2.imshow("original",sample)
    # cv2.imshow("Original", cv2.resize(sample, None, fx=1, fy=1))
    # cv2.waitKey(0)
    # cv2.destroyAllWindows()

    best_score = counter = 0
    filename = image = kp1 = kp2 = mp = None
    for file in os.listdir(path_to_dataset):
        fingerprint_img = cv2.imread(f'{path_to_dataset}/' + file)
    #   fingerprint_img=abs(140-fingerprint_img)
    #   fingerprint_img=fingerprint_img[90:310 , 170:480]

        sift = (cv2.SIFT_create())
        keypoints_1, des1 = sift.detectAndCompute(sample, None)
        keypoints_2, des2 = sift.detectAndCompute(fingerprint_img, None)
    #   index_params = dict(algorithm = FLANN_INDEX_KDTREE, trees = 5)

        matches = cv2.FlannBasedMatcher(
            {"algorithm": 1, "trees": 10}, dict(checks=50)).knnMatch(des1, des2, k=2)
    #   print(matches)

        match_points = []
        for p, q in matches:
            if p.distance < 0.1 * q.distance:   # it is lowe's ratio test checks if two distances are sufficiently different
                match_points.append(p)

        keypoints = 0
        if len(keypoints_1) <= len(keypoints_2):
            keypoints = len(keypoints_1)
        else:
            keypoints = len(keypoints_2)

        if len(match_points) / keypoints * 100 > best_score:
            best_score = len(match_points) / keypoints * 100
            filename = file
            image = fingerprint_img
            kp1, kp2, mp = keypoints_1, keypoints_2, match_points

        if best_score > 60:  
            break


    if best_score < 40:
        print("Not matched")
        return False
    else:
        print("Best match:  " + filename)
        print("Best score:  " + str(best_score))
        result = cv2.drawMatches(sample, kp1, image, kp2, mp, None)
        result = cv2.resize(result, None, fx=1, fy=1)
        # cv2.imshow("Result", result)
        # cv2.waitKey(0)
        # cv2.destroyAllWindows()
        return True







try:
    #Reader.write("Ansh Chawla")
    id , text = reader.read()
    print(id)
    print(text)
    row = sheet.max_row+1

    person = read(id)

    
    
    if(person == None):
        raise IndexError("person not found invalid uid")
    else:
        #print(sheet[f'A{row}:F{row}'][0][1].value)
        if (direction == GOINGOUT and person["location"] == "out") or (direction == GOINGIN and person["location"] == "in"):
            sheet[f'{mapping["remarks"]}{row}']= "Card RESCAN"
        sheet[f'{mapping["time"]}{row}']= datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        sheet[f'{mapping["name"]}{row}']= person["name"]
        sheet[f'{mapping["sid"]}{row}']= person["sid"]
        sheet[f'{mapping["phone"]}{row}']= person["phone"]
        sheet[f'{mapping["uid"]}{row}']= id

        if direction == GOINGOUT:
            sheet[f'{mapping["direction"]}{row}'] = "EXIT" # print name
            person["location"] = "out"
            write(person)
            time.sleep(2)
        elif direction == GOINGIN:
            sheet[f'{mapping["direction"]}{row}'] = "ENTRY" 
            person["location"] = "in"
            write(person)
            
    matched = False
    chance = 5
    if(direction == GOINGIN):
        while not matched and chance != 0:
            GPIO.output(led,GPIO.HIGH)
            if not GPIO.input(touch):
                time.sleep(1)
                chance = chance-1
                print("given chance")
                image_path = "image.jpg"  #address and name of image
                camera.capture(image_path) 
                image = cv2.imread(image_path)
                img = cv2.cvtColor(image,cv2.COLOR_BGR2GRAY)
                cv2.imwrite("image_grayscale.jpg",img)
                # call image_check

                GPIO.output(led,GPIO.LOW)
                matched = match_finger(image)
                
                # if cv2.waitKey(0) == ord('s'):
                #     break
        if not matched:
            print("sry no match") 
            sheet[f'{mapping["remarks"]}{row}']= "FINGERPRINT MISMATCH"
                
    else:
        # open the gate
        pass


finally:
    GPIO.cleanup()
    workbook.save(filename="attendance.xlsx")
    dataset.save(path_to_database)