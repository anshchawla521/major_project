import schedule
import time
import twilio_sms
import openpyxl as op

path_to_dataset = "./images"
dataset = op.load_workbook(path_to_database)
sheet_dataset = dataset.active

workbook = op.load_workbook(filename="attendance.xlsx")
sheet = workbook["Sheet1"]

phone = None
name = None

def get_col(col_name):
    col = 1
    while col <= sheet.max_column:
        if sheet.cell(row = 1, column= col).value == col_name:
            return col
        col += 1
    return -1

def job():
    col_no = get_col("Location")
    if col_no == -1:
        return "Not Found"
    else:
        present = 0
        missing = 0
        i = 2
        while i <= sheet.max_row:
            if sheet.cell(row = i, column=col_no).value == "in":
                present += 1
            elif sheet.cell(row = i, column=col_no).value == "out":
                missing += 1
            i += 1
        twilio_sms.sms_auth(str(present),str(missing), str(phone_number), str(name))

schedule.every().day.at("21:00").do(job)